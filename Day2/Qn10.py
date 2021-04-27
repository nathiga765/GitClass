import keyboard
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
import openpyxl

driver = webdriver.Chrome(
    executable_path=r"C:\Users\AIT-LP09\PycharmProjects\py37\Driver\chromedriver_win32\chromedriver.exe")
driver.get("https://www.amazon.in/")

driver.maximize_window()
btn_search = driver.find_element(By.ID, "twotabsearchtextbox")
btn_search.send_keys("iphone", Keys.ENTER)
product = driver.find_elements(By.XPATH, "//span[@class='a-size-medium a-color-base a-text-normal']")
price = driver.find_elements(By.XPATH, "//span[@class='a-price-whole']")
d = {}
print(len(price))
print(len(product))
for i in range(0, len(product) - 1):
    a = int(price[i].text.replace(',', ''))
    d.update({a: product[i].text})
print(d)
l = sorted(d)
exc_loc = r"C:\Users\AIT-LP09\Documents\Outlook Files\Documents\amazon.xlsx"
w = openpyxl.Workbook()
s = w.create_sheet("amount", 0)
for x in range(1, len(l)):
    c = s.cell(x, 1)
    c.value = l[x]
    q = s.cell(x, 2)
    q.value = d.get(l[x])
    w.save(exc_loc)






# print(price)
# print(product)


# btn_watch = driver.find_element(By.XPATH,"//div[@data-keyword='iphone 11']").click()
#
# keyboard.press("enter")
# keyboard.release("enter")
# keyboard.press("down")
# keyboard.release("down")
# keyboard.press("down")
# keyboard.release("down")
# keyboard.press("down")
# keyboard.release("down")
# keyboard.press("enter")
# keyboard.release("enter")
